from selenium import webdriver
from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
import datetime
import pymysql
import time


# 로그인
def fall_login():
    driver.get("https://fall-mvp.com/")
    driver.find_element_by_name('userid').send_keys("kosae08")
    driver.find_element_by_name('userpass').send_keys("Joker0422")
    driver.find_element_by_xpath('//*[@id="loginForm1"]/button').click()
    time.sleep(1)

    if driver.current_url == "https://fall-mvp.com/main":
        return True
    else:
        return False


# 할당량 확인 시 return True, 아니면 False
def today_quota_check():
    sheet_wb = load_workbook("September.xlsx", data_only=True)
    sheet_ws = sheet_wb['달력']
    date = datetime.datetime.now()
    now = date.strftime("%Y-%m-%d")

    for row_count in range(5, 13, 1):
        for col_count in range(1, 7):
            if str(now) in str(sheet_ws.cell(row_count, col_count).value):
                today_row_count, today_col_count = row_count, col_count

    if sheet_ws.cell(today_row_count+1, today_col_count).value is None:
        return False
    else:
        return True


# 1.2 ~ 1.4 배당 확인
def fall_filtering():
    id_list = []
    team_list, odds_list, match_list = [], [], []
    driver.get("https://fall-mvp.com/main/cross.html")
    r = driver.page_source
    soup = BeautifulSoup(r, 'html.parser')

    for top in soup.find_all('ul', class_='g_item'):
        for mid in top.find_all('li'):
            for bot in mid.find_all('span', class_='g_home_odd_n'):
                if "," in str(bot.get_text().strip()):
                    if 1.2 < float(str(bot.get_text().strip()).replace(",", "")) < 1.4:
                        odds_list.append(str(bot.get_text().strip()).replace(",", ""))
                        id_list.append(mid.get('id'))
                        for team in mid.find_all('span', class_='g_away_n'):
                            team_list.append(team.get_text().strip())
                        for match in top.find_all('li', class_='g_day'):
                            match_list.append("2021-" + str(match.get_text().strip()))
                else:
                    if 1.2 < float(str(bot.get_text().strip()).replace(",", "")) < 1.4:
                        odds_list.append(str(bot.get_text().strip()).replace(",", ""))
                        id_list.append(mid.get('id'))
                        for team in mid.find_all('span', class_='g_home_n'):
                            team_list.append(team.get_text().strip())
                        for match in top.find_all('li', class_='g_day'):
                            match_list.append("2021-" + str(match.get_text().strip()))
            for bot in mid.find_all('span', class_='g_away_odd_n'):
                if "," in str(bot.get_text().strip()):
                    if 1.2 < float(str(bot.get_text().strip()).replace(",", "")) < 1.4:
                        odds_list.append(str(bot.get_text().strip()).replace(",", ""))
                        id_list.append(mid.get('id'))
                        for team in mid.find_all('span', class_='g_away_n'):
                            team_list.append(team.get_text().strip())
                        for match in top.find_all('li', class_='g_day'):
                            match_list.append("2021-" + str(match.get_text().strip()))
                else:
                    if 1.2 < float(str(bot.get_text().strip()).replace(",", "")) < 1.4:
                        odds_list.append(str(bot.get_text().strip()).replace(",", ""))
                        id_list.append(mid.get('id'))
                        for team in mid.find_all('span', class_='g_away_n'):
                            team_list.append(team.get_text().strip())
                        for match in top.find_all('li', class_='g_day'):
                            match_list.append("2021-" + str(match.get_text().strip()))

    match_list_input(id_list, team_list, odds_list, match_list)


# 엑셀 파일에 저장
# user_data1 = 아이디, user_data2 = 팀 이름, user_data3 = 팀 배당, user_data4 = 경기 시간
def match_list_input(user_data1, user_data2, user_data3, user_data4):
    user_list_1 = []

    sql = "select team from match_list"
    cursor.execute(sql)
    conn.commit()

    datas = cursor.fetchall()

    for team in datas:
        user_list_1.append(team)

    for match, source_id, team, odds in zip(user_data4, user_data1, user_data2, user_data3):
        if team not in user_list_1:
            sql = "insert into match_list (match_time, source_id, team, odds) values (%s, %s, %s, %s)"
            values = (match, source_id, team, odds)
            cursor.execute(sql, values)
            conn.commit()


# 경기 목록 리스트업 후 사용자의 입력을 받아 베팅 진행
def betting_listup():
    betting_flag = False
    no_list, time_list, source_id_list, team_list, odds_list = [], [], [], [], []

    sql = "select no, match_time, team, odds from match_list"
    cursor.execute(sql)

    datas = cursor.fetchall()

    for no, match, team, odds in datas:
        no_list.append(no)
        time_list.append(match)
        team_list.append(team)
        odds_list.append(odds)

    for no, match, team, odds in zip(no_list, time_list, team_list, odds_list):
        difference = match-datetime.datetime.now()

        if int(difference.days) == 0:
            print(str(no) + "      " + str(match) + " " + str(team) + " " + str(odds))

    input_match_no = int(input("경기를 골라주세요: "))
    input_account = input("금액을 입력해주세요: ")

    sql = "select source_id, team, odds from match_list where no=" + "'" + str(input_match_no) + "'"
    cursor.execute(sql)
    conn.commit()

    datas = cursor.fetchall()
    # 입력받은 경기, 베팅금액을 DB에 저장
    for source_id, team, odds in datas:
        if driver.find_element_by_id(source_id):
            driver.find_element_by_id(source_id).click()
            driver.find_element_by_id('nAmt').send_keys(input_account)
            driver.find_element_by_xpath('//*[@id="cart"]/div[3]/button').send_keys(Keys.ENTER)
            time.sleep(1)
            alert = Alert(driver)
            alert.accept()
            time.sleep(3)
            alert = Alert(driver)
            if "확인" in str(alert.text):
                alert.accept()
                betting_content_insert(team, odds, input_account)
            else:
                alert.accept()
                print('다시 확인해주세요')
        else:
            pass


# DB 연동 완료(테스트 필요), 모든 데이터 대상(빅데이터 관리)
def results_check():
    win_list = []

    page_count = 0

    url = "https://fall-mvp.com/main/result/D/"
    url_a = ".html"

    while page_count < 101:
        com_url = url + str(page_count) + url_a

        driver.get(com_url)
        r = driver.page_source
        soup = BeautifulSoup(r, 'html.parser')

        for top in soup.find_all('li', class_='g_home_ed g_gr_o'):
            for mid in top.find_all('span', class_='g_home_o'):
                win_list.append(mid.get_text().strip())
        for top in soup.find_all('li', class_='g_away_ed g_gr_o'):
            for mid in top.find_all('span', class_='g_away_o'):
                win_list.append(mid.get_text().strip())
        page_count += 25

    for win_team in win_list:
        sql = "update match_list set result=True where team=" + "'" + str(win_team) + "'"
        cursor.execute(sql)
        conn.commit()

    betting_content_modify()


# user_data1 = 팀 이름, user_data2 = 배당률, user_data3 = 베팅 금액
def betting_content_insert(user_data1, user_data2, user_data3):
    sql = "insert into betting_data (team, odds, amount) values (%s, %s, %s)"
    values = user_data1, user_data2, user_data3

    cursor.execute(sql, values)
    conn.commit()


def betting_content_modify():
    sql = "select no, team, odds, amount, result from betting_data"
    cursor.execute(sql)
    conn.commit()

    datas = cursor.fetchall()

    for no, team, odds, amount, result in datas:
        if result is None:
            sql = "select result, odds from match_list where team=" + "'" + team + "'"
            cursor.execute(sql)
            conn.commit()

            datas_1 = cursor.fetchall()
            for result in datas_1:
                if result:
                    sql = "select total_plus from betting_data where no=" + "'" + str(int(no) - 1) + "'"
                    cursor.execute(sql)
                    conn.commit()
                    datas_2 = cursor.fetchall()

                    for total_plus in datas_2:
                        total_plus = str(total_plus).replace("(", "").replace(",", "").replace(")", "")

                        sql = "update betting_data set result=True, plus=" + "'" + str(amount * odds - amount) + "'" + \
                              ",total_plus=" + "'" + str(
                            int(total_plus) + amount * odds - amount) + "'" + "where team=" + "'" + str(team) + "'"
                        cursor.execute(sql)
                        conn.commit()


if __name__ == '__main__':
    driver = webdriver.Chrome(ChromeDriverManager().install())
    conn = pymysql.connect(host='localhost', user='root', password="Joker0422!", db="project")
    cursor = conn.cursor()
    mode = int(input("모드를 선택하세요: (1. 베팅, 2. 빅데이터)"))

    while True:
        if fall_login():
            fall_filtering()
            if mode is 1:
                betting_listup()
            else:
                results_check()
        time.sleep(3600)


# 베팅 전 목록 리스트업 과정에서 현재 시간보다 지난 경기 제외
